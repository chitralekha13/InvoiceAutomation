"""
Shared helper functions for Azure Functions
Provides utilities for SharePoint, SQL, authentication, and logging
"""
import os
import json
import logging
import tempfile
import base64
import psycopg2
from psycopg2.extras import RealDictCursor
import jwt
import re
from datetime import datetime
from typing import Dict, Optional, Any
from office365.sharepoint.client_context import ClientContext
from office365.runtime.auth.client_credential import ClientCredential

logger = logging.getLogger(__name__)

# ============================================================================
# SharePoint Helpers
# ============================================================================

def _sharepoint_site_server_relative_prefix() -> str:
    """
    Return the server-relative prefix for the configured SharePoint site URL.
    Example:
      SHAREPOINT_SITE_URL = https://tenant.sharepoint.com/sites/Accounts
      -> '/sites/Accounts'
    """
    site_url = os.environ.get("SHAREPOINT_SITE_URL") or ""
    try:
        from urllib.parse import urlparse
        parsed = urlparse(site_url)
        path = (parsed.path or "").rstrip("/")
        return path or ""
    except Exception:
        return ""

def _normalize_server_relative_url(path: str) -> str:
    """
    Normalize a SharePoint path to a server-relative URL suitable for
    ctx.web.get_file_by_server_relative_url().

    Accepts:
    - Full server-relative: '/sites/Accounts/Invoices/File.xlsx'
    - Site-relative with leading slash: '/Invoices/File.xlsx'
    - Library-relative: 'Invoices/File.xlsx'
    """
    p = (path or "").strip().replace("\\", "/")
    if not p:
        raise ValueError("Empty SharePoint path")

    # Already server-relative to a site/web
    if p.startswith("/sites/") or p.startswith("/teams/"):
        return p

    prefix = _sharepoint_site_server_relative_prefix()
    if not prefix:
        # Best-effort fallback; still allow absolute-ish paths.
        return p if p.startswith("/") else f"/{p}"

    # If caller passed '/Invoices/..' treat as site-relative
    if p.startswith("/"):
        return f"{prefix}{p}"
    return f"{prefix}/{p}"

def get_sharepoint_excel_url() -> Optional[str]:
    """Return the full SharePoint URL for the Excel file (for Download Excel button)."""
    site_url = (os.environ.get("SHAREPOINT_SITE_URL") or "").rstrip("/")
    excel_path = os.environ.get("SHAREPOINT_EXCEL_PATH") or "Invoices/Invoice_Register_Master.xlsx"
    if not site_url:
        return None
    try:
        from urllib.parse import urlparse
        parsed = urlparse(site_url)
        base = f"{parsed.scheme}://{parsed.netloc}"
        server_path = _normalize_server_relative_url(excel_path)
        return base + server_path
    except Exception:
        return None


def _get_sharepoint_tenant(site_url: str) -> str:
    """Extract tenant identifier from SharePoint site URL for certificate auth."""
    tenant_name = os.environ.get('SHAREPOINT_TENANT_NAME')
    if tenant_name:
        return tenant_name
    tenant_id = os.environ.get('AZURE_TENANT_ID')
    if tenant_id:
        return tenant_id
    # Derive from hostname: invoiveautomation.sharepoint.com -> invoiveautomation.onmicrosoft.com
    try:
        from urllib.parse import urlparse
        host = urlparse(site_url).netloc
        if '.sharepoint.com' in host:
            subdomain = host.split('.sharepoint.com')[0]
            return f"{subdomain}.onmicrosoft.com"
    except Exception:
        pass
    raise ValueError(
        "Set SHAREPOINT_TENANT_NAME (e.g. invoiveautomation.onmicrosoft.com) or "
        "AZURE_TENANT_ID for SharePoint certificate authentication"
    )

def get_sharepoint_context() -> ClientContext:
    """
    Get authenticated SharePoint context.
    Uses certificate auth (SHAREPOINT_CERT_BASE64 + SHAREPOINT_CERT_THUMBPRINT) if set,
    otherwise falls back to client secret (ClientCredential) - note: client secret
    does NOT work with Azure AD app-only for SharePoint REST API; use certificate.
    """
    site_url = os.environ.get('SHAREPOINT_SITE_URL')
    client_id = os.environ.get('AZURE_CLIENT_ID')

    if not site_url or not client_id:
        raise ValueError("SHAREPOINT_SITE_URL and AZURE_CLIENT_ID are required")

    cert_base64 = os.environ.get('SHAREPOINT_CERT_BASE64')
    cert_thumbprint = os.environ.get('SHAREPOINT_CERT_THUMBPRINT')

    if cert_base64 and cert_thumbprint:
        # Certificate-based auth (required for Azure AD app-only with SharePoint REST API)
        tenant = _get_sharepoint_tenant(site_url)
        try:
            pem_bytes = base64.b64decode(cert_base64)
            pem_content = pem_bytes.decode('utf-8') if isinstance(pem_bytes, bytes) else str(pem_bytes)
        except Exception as e:
            raise ValueError(f"Invalid SHAREPOINT_CERT_BASE64: {e}") from e

        fd, cert_path = tempfile.mkstemp(suffix='.pem')
        try:
            os.write(fd, pem_content.encode('utf-8') if isinstance(pem_content, str) else pem_content)
            os.close(fd)
        except Exception:
            try:
                os.unlink(cert_path)
            except OSError:
                pass
            raise

        cert_settings = {
            'client_id': client_id,
            'thumbprint': cert_thumbprint.strip(),
            'cert_path': cert_path,
        }
        return ClientContext(site_url).with_client_certificate(tenant, **cert_settings)
    else:
        # Client secret (only works with deprecated SharePoint Add-in, not Azure AD app)
        client_secret = os.environ.get('AZURE_CLIENT_SECRET')
        if not client_secret:
            raise ValueError(
                "SharePoint certificate auth requires SHAREPOINT_CERT_BASE64 and "
                "SHAREPOINT_CERT_THUMBPRINT. Client secret (Azure AD app) does not work "
                "for SharePoint REST API app-only access."
            )
        credentials = ClientCredential(client_id, client_secret)
        return ClientContext(site_url).with_credentials(credentials)

def upload_file_to_sharepoint(file_content: bytes, file_name: str, folder_path: str = "Invoices") -> str:
    """
    Upload file to SharePoint document library.
    
    Args:
        file_content: File content as bytes
        file_name: Name of the file
        folder_path: Library name (e.g. 'Invoices' or 'JSON_Logs') or path like 'Invoices/2025/01_January'
    
    Returns:
        Server-relative URL of uploaded file
    """
    ctx = get_sharepoint_context()
    parts = [p for p in folder_path.replace("\\", "/").strip("/").split("/") if p.strip()]
    if not parts:
        parts = ["Invoices"]
    list_name = parts[0]
    root = ctx.web.lists.get_by_title(list_name).root_folder
    root.get().execute_query()
    target_folder = root
    for part in parts[1:]:
        try:
            target_folder = target_folder.folders.get_by_name(part).get().execute_query()
        except Exception:
            target_folder = target_folder.folders.add(part).execute_query()
    uploaded_file = target_folder.upload_file(file_name, file_content).execute_query()
    return uploaded_file.properties.get("ServerRelativeUrl") or uploaded_file.properties.get("serverRelativeUrl") or ""

def download_file_from_sharepoint(file_path: str) -> bytes:
    """Download file from SharePoint"""
    ctx = get_sharepoint_context()
    file = ctx.web.get_file_by_server_relative_url(file_path)
    file_content = file.read()
    return file_content

def save_json_to_sharepoint(json_data: Dict, file_name: str, folder_path: str = 'JSON files') -> str:
    """Save JSON data to SharePoint JSON files library"""
    now = datetime.now()
    year = now.strftime('%Y')
    month = now.strftime('%m_%B')
    full_folder_path = f'{folder_path}/{year}/{month}'
    
    json_content = json.dumps(json_data, indent=2).encode('utf-8')
    return upload_file_to_sharepoint(json_content, file_name, full_folder_path)

# ============================================================================
# PostgreSQL Database Helpers
# ============================================================================

def get_sql_connection():
    """Get PostgreSQL database connection"""
    conn_str = os.environ.get('SQL_CONNECTION_STRING')
    if not conn_str:
        raise ValueError("SQL_CONNECTION_STRING not found in environment")
    return psycopg2.connect(conn_str)

def insert_invoice(invoice_id: str, vendor_id: str, doc_name: str, pdf_url: str, **kwargs) -> None:
    """Insert new invoice record into PostgreSQL database"""
    conn = get_sql_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            INSERT INTO invoices (
                invoice_id, vendor_id, doc_name, pdf_url, status, created_at, invoice_received_date
            ) VALUES (%s, %s, %s, %s, 'Pending', NOW(), NOW())
        """, (invoice_id, vendor_id, doc_name, pdf_url))
        
        conn.commit()
        logger.info(f"Inserted invoice {invoice_id} into PostgreSQL database")
    finally:
        cursor.close()
        conn.close()


def update_invoice(invoice_id: str, **kwargs) -> None:
    """Update invoice record in PostgreSQL database"""
    conn = get_sql_connection()
    cursor = conn.cursor()
    
    try:
        # Build dynamic UPDATE query based on provided kwargs
        if not kwargs:
            return
        
        set_clauses = []
        values = []
        param_num = 1
        
        for key, value in kwargs.items():
            set_clauses.append(f"{key} = %s")
            values.append(value)
            param_num += 1
        
        set_clauses.append("last_updated_at = NOW()")
        values.append(invoice_id)
        
        query = f"""
            UPDATE invoices
            SET {', '.join(set_clauses)}
            WHERE invoice_id = %s
        """
        
        cursor.execute(query, values)
        conn.commit()
        logger.info(f"Updated invoice {invoice_id} in PostgreSQL database")
    finally:
        cursor.close()
        conn.close()

def find_duplicate_invoice(fields: Dict) -> Optional[str]:
    """
    Check if an existing row matches all key fields. Returns existing invoice_id if duplicate, else None.
    Date is the main differentiator: same vendor/amount but different month = NOT duplicate.
    Key fields: invoice_number, vendor_name, invoice_amount, invoice_date (or start_date).
    """
    inv_num = (fields.get("invoice_number") or "").strip()
    vendor = (str(fields.get("vendor_name") or "").strip()).lower()
    amount = fields.get("invoice_amount")
    inv_date = fields.get("invoice_date") or fields.get("start_date") or fields.get("end_date")
    hours = fields.get("invoice_hours") or fields.get("vendor_hours")
    if not inv_num and not (vendor and amount and inv_date):
        return None
    conn = get_sql_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    try:
        if inv_num:
            cursor.execute(
                "SELECT invoice_id, invoice_number, vendor_name, invoice_amount, invoice_date, start_date, end_date, invoice_hours FROM invoices WHERE LOWER(TRIM(invoice_number)) = LOWER(%s)",
                (inv_num.strip(),)
            )
        else:
            cursor.execute(
                "SELECT invoice_id, invoice_number, vendor_name, invoice_amount, invoice_date, start_date, end_date, invoice_hours FROM invoices"
            )
        rows = cursor.fetchall()
        new_date = str(inv_date or "").strip()[:10] if inv_date else None
        for row in rows:
            r_num = (row.get("invoice_number") or "").strip()
            r_vendor = (str(row.get("vendor_name") or "").strip()).lower()
            r_amount = row.get("invoice_amount")
            r_date = row.get("invoice_date") or row.get("start_date") or row.get("end_date")
            r_date_str = str(r_date or "").strip()[:10] if r_date else None
            r_hours = row.get("invoice_hours")
            if inv_num and r_num.lower() != inv_num.lower():
                continue
            if vendor and r_vendor != vendor:
                continue
            if amount is not None:
                try:
                    if abs(float(r_amount or 0) - float(amount)) > 0.01:
                        continue
                except (TypeError, ValueError):
                    continue
            if new_date and r_date_str:
                if new_date != r_date_str:
                    continue
            elif new_date or r_date_str:
                continue
            if hours is not None:
                try:
                    if abs(float(r_hours or 0) - float(hours)) > 0.01:
                        continue
                except (TypeError, ValueError):
                    pass
            return str(row.get("invoice_id"))
        return None
    finally:
        cursor.close()
        conn.close()


def delete_invoice(invoice_id: str) -> bool:
    """Delete invoice from PostgreSQL. Returns True if deleted, False if not found."""
    conn = get_sql_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM invoices WHERE invoice_id = %s", (invoice_id,))
        conn.commit()
        deleted = cursor.rowcount > 0
        if deleted:
            logger.info("Deleted invoice %s from database", invoice_id)
        return deleted
    finally:
        cursor.close()
        conn.close()


def get_invoice(invoice_id: str) -> Optional[Dict]:
    """Get invoice record from PostgreSQL database"""
    conn = get_sql_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cursor.execute("SELECT * FROM invoices WHERE invoice_id = %s", (invoice_id,))
        row = cursor.fetchone()
        
        if not row:
            return None
        
        invoice = dict(row)
        
        # Convert datetime objects to ISO strings
        for key, value in invoice.items():
            if hasattr(value, 'isoformat'):
                invoice[key] = value.isoformat()
        
        return invoice
    finally:
        cursor.close()
        conn.close()

def get_invoices_by_vendor(vendor_id: str) -> list:
    """Get all invoices for a vendor from PostgreSQL"""
    conn = get_sql_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cursor.execute(
            "SELECT * FROM invoices WHERE vendor_id = %s ORDER BY created_at DESC",
            (vendor_id,)
        )
        
        rows = cursor.fetchall()
        
        invoices = []
        for row in rows:
            invoice = dict(row)
            # Convert datetime objects
            for key, value in invoice.items():
                if hasattr(value, 'isoformat'):
                    invoice[key] = value.isoformat()
            invoices.append(invoice)
        
        return invoices
    finally:
        cursor.close()
        conn.close()

def get_all_invoices() -> list:
    """Get all invoices (for accounts team) from PostgreSQL"""
    conn = get_sql_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cursor.execute("SELECT * FROM invoices ORDER BY created_at DESC")
        
        rows = cursor.fetchall()
        
        invoices = []
        for row in rows:
            invoice = dict(row)
            # Convert datetime objects
            for key, value in invoice.items():
                if hasattr(value, 'isoformat'):
                    invoice[key] = value.isoformat()
            invoices.append(invoice)
        
        return invoices
    finally:
        cursor.close()
        conn.close()

# ============================================================================
# Authentication Helpers
# ============================================================================

def extract_token_from_request(req) -> Optional[str]:
    """Extract JWT token from request Authorization header"""
    auth_header = req.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return None
    return auth_header.split(' ')[1]

def decode_token(token: str) -> Dict:
    """Decode JWT token (without verification for now)"""
    try:
        # In production, verify the token signature
        decoded = jwt.decode(token, options={"verify_signature": False})
        return decoded
    except Exception as e:
        logger.error(f"Failed to decode token: {str(e)}")
        raise

def extract_vendor_id_from_token(token: str) -> str:
    """Extract vendor_id from JWT token"""
    decoded = decode_token(token)
    # Try different possible fields
    return (
        decoded.get('email') or
        decoded.get('upn') or
        decoded.get('preferred_username') or
        decoded.get('sub') or
        decoded.get('oid')
    )

def extract_user_id_from_token(token: str) -> str:
    """Extract user identifier from JWT token"""
    decoded = decode_token(token)
    return decoded.get('email') or decoded.get('upn') or decoded.get('preferred_username')

def check_manager_permission(token: str) -> bool:
    """Check if user has manager/approver role"""
    decoded = decode_token(token)
    roles = decoded.get('roles', [])
    # Check for manager role (configure in Azure AD app registration)
    return 'Invoice.Approver' in roles or 'Manager' in roles or 'Admin' in roles


def _row_to_dashboard(row: Dict) -> Dict:
    """Map SQL row to dashboard row format (invoice_uuid, pay_period_start, net_terms, etc.)."""
    out = dict(row)
    if "invoice_id" in out and "invoice_uuid" not in out:
        out["invoice_uuid"] = str(out["invoice_id"])
    # Dashboard uses pay_period_start/end, net_terms, pay_rate, vendor_hours
    if out.get("start_date") and "pay_period_start" not in out:
        out["pay_period_start"] = out["start_date"]
    if out.get("end_date") and "pay_period_end" not in out:
        out["pay_period_end"] = out["end_date"]
    if out.get("payment_terms") and "net_terms" not in out:
        out["net_terms"] = out["payment_terms"]
    if out.get("hourly_rate") is not None and "pay_rate" not in out:
        out["pay_rate"] = out["hourly_rate"]
    if out.get("invoice_hours") is not None and "vendor_hours" not in out:
        out["vendor_hours"] = out["invoice_hours"]
    status = out.get("approval_status") or out.get("status") or "Pending"
    if status in ("To Start", "In Progress"):
        status = "Pending"
    if status == "Need Approval":
        status = "NEED APPROVAL"
    out["approval_status"] = status
    if "status" not in out or out["status"] is None:
        out["status"] = status
    if "orchestrator_summary" not in out or out["orchestrator_summary"] is None:
        out["orchestrator_summary"] = out.get("last_agent_text") or ""
    return out


def _dashboard_metrics(rows: list) -> Dict:
    """Compute dashboard metrics from rows."""
    total = len(rows)
    pending = sum(1 for r in rows if (r.get("approval_status") or r.get("status")) == "Pending")
    complete = sum(1 for r in rows if (r.get("approval_status") or r.get("status")) == "Complete")
    need_approval = sum(1 for r in rows if (r.get("approval_status") or r.get("status")) == "NEED APPROVAL")
    payment_initiated = sum(1 for r in rows if r.get("bill_pay_initiated_on"))
    total_amount = 0.0
    for r in rows:
        try:
            v = r.get("invoice_amount")
            if v is not None:
                total_amount += float(v)
        except (TypeError, ValueError):
            pass
    return {
        "total": total,
        "pending": pending,
        "complete": complete,
        "need_approval": need_approval,
        "payment_initiated": payment_initiated,
        "total_amount": round(total_amount, 2),
    }


def get_dashboard_payload(req) -> tuple:
    """
    Get dashboard data (rows + metrics) based on request token.
    Returns (list of dashboard rows, metrics dict).
    When SQL_CONNECTION_STRING is not set, returns empty rows and zero metrics.
    """
    if not os.environ.get('SQL_CONNECTION_STRING'):
        return [], {
            "total": 0, "pending": 0, "complete": 0, "need_approval": 0,
            "payment_initiated": 0, "total_amount": 0.0,
        }
    token = extract_token_from_request(req)
    is_manager = False
    vendor_id = None
    if token:
        try:
            is_manager = check_manager_permission(token)
            vendor_id = extract_vendor_id_from_token(token)
        except Exception:
            pass
    if is_manager or not vendor_id:
        rows = get_all_invoices()
    else:
        rows = get_invoices_by_vendor(vendor_id)
    dashboard_rows = [_row_to_dashboard(r) for r in rows]
    metrics = _dashboard_metrics(dashboard_rows)
    return dashboard_rows, metrics


# ============================================================================
# Document Intelligence Helpers
# ============================================================================

def analyze_invoice_bytes(file_content: bytes, filename: str = "invoice.pdf") -> Optional[Dict]:
    """
    Analyze invoice PDF/image bytes with Azure Document Intelligence (prebuilt-invoice).
    Returns full_text, extracted_text for iGentic, plus structured_fields (InvoiceId, VendorName, etc.).
    """
    import requests
    import time
    endpoint = os.environ.get('AZURE_DI_ENDPOINT', '').rstrip('/')
    key = os.environ.get('AZURE_DI_KEY')
    if not endpoint or not key:
        logger.info("Document Intelligence skipped: AZURE_DI_ENDPOINT or AZURE_DI_KEY not set")
        return None

    content_type = "application/pdf"
    if filename.lower().endswith(('.png', '.jpg', '.jpeg')):
        content_type = "image/jpeg" if 'jpg' in filename.lower() or 'jpeg' in filename.lower() else "image/png"
    headers = {"Ocp-Apim-Subscription-Key": key, "Content-Type": content_type}

    for path_prefix, api_version in [("formrecognizer", "2023-07-31"), ("documentintelligence", "2023-07-31")]:
        analyze_url = f"{endpoint}/{path_prefix}/documentModels/prebuilt-invoice:analyze?api-version={api_version}"
        try:
            resp = requests.post(analyze_url, headers=headers, data=file_content, timeout=60)
            if resp.status_code in (404, 400, 401):
                continue
            if resp.status_code not in (200, 202):
                logger.warning("Document Intelligence analyze failed: HTTP %s", resp.status_code)
                continue
            operation_location = resp.headers.get("Operation-Location")
            if not operation_location:
                continue

            for _ in range(60):
                time.sleep(1)
                poll_resp = requests.get(operation_location, headers={"Ocp-Apim-Subscription-Key": key}, timeout=30)
                if poll_resp.status_code != 200:
                    continue
                result = poll_resp.json()
                if result.get("status") == "succeeded":
                    payload = result.get("analyzeResult") or result.get("result") or result
                    extracted_text = []
                    full_text_parts = []

                    for page in (payload.get("pages") or []):
                        for line in (page.get("lines") or []):
                            t = (line.get("content") or "").strip()
                            extracted_text.append(t)
                            full_text_parts.append(t)
                    full_text = "\n".join(full_text_parts) if full_text_parts else (payload.get("content") or "")
                    if not full_text and payload.get("content"):
                        full_text = payload["content"]
                        extracted_text = [s.strip() for s in full_text.split("\n") if s.strip()]

                    structured_fields = _extract_invoice_fields(payload)

                    logger.info("Document Intelligence succeeded for %s: %d lines, %d structured fields",
                                filename, len(extracted_text), len(structured_fields))

                    return {
                        "timestamp": datetime.now().isoformat(),
                        "file_path": filename,
                        "extracted_text": extracted_text,
                        "full_text": full_text[:15000] if full_text else "",
                        "structured_fields": structured_fields,
                        "status": "success",
                        "source": "document_intelligence_rest",
                    }
                if result.get("status") == "failed":
                    logger.warning("Document Intelligence analysis failed: %s", result.get("error", {}))
                    break
        except Exception as e:
            logger.warning("Document Intelligence %s error: %s", path_prefix, e)
            continue
    return None


def _extract_invoice_fields(payload: Dict) -> Dict:
    """Extract structured invoice fields from Document Intelligence prebuilt-invoice result."""
    out = {}
    docs = payload.get("documents") or []
    if not docs:
        return out
    fields = docs[0].get("fields") or {}
    field_map = {
        "InvoiceId": "invoice_number",
        "VendorName": "vendor_name",
        "VendorAddress": "vendor_address",
        "InvoiceDate": "invoice_date",
        "DueDate": "due_date",
        "InvoiceTotal": "invoice_total",
        "AmountDue": "amount_due",
        "SubTotal": "sub_total",
        "TotalTax": "total_tax",
        "CustomerName": "customer_name",
        "PurchaseOrder": "purchase_order",
        "PaymentTerm": "payment_term",
    }
    for api_key, our_key in field_map.items():
        obj = fields.get(api_key)
        if not isinstance(obj, dict):
            continue
        val = obj.get("value")
        if val is not None:
            if hasattr(val, "isoformat"):
                val = val.isoformat()
            out[our_key] = val
    # Extract hours from line items (Quantity) - sum for consulting/time-based invoices
    items = docs[0].get("items") or []
    if not items and isinstance(fields.get("Items"), dict):
        items = fields["Items"].get("value") or fields["Items"].get("valueArray") or []
    if isinstance(items, list):
        total_qty = 0
        for item in items:
            if isinstance(item, dict):
                qty_obj = item.get("Quantity") or item.get("quantity")
                if isinstance(qty_obj, dict):
                    qty_val = qty_obj.get("value") or qty_obj.get("content")
                    try:
                        total_qty += float(qty_val or 0)
                    except (TypeError, ValueError):
                        pass
                elif isinstance(qty_obj, (int, float)):
                    total_qty += float(qty_obj)
        if total_qty > 0:
            out["invoice_hours"] = total_qty
    return out


def _parse_hours_from_text(text: str) -> Optional[float]:
    """Try to extract hours from invoice full_text. Returns first plausible match."""
    if not text or not isinstance(text, str):
        return None
    text_lower = text.lower()
    patterns = [
        r'(?:total\s+)?(?:billable\s+)?(?:invoice\s+)?hours\s*[:\-]?\s*(\d+(?:\.\d+)?)',
        r'(\d+(?:\.\d+)?)\s*(?:hours?)',
        r'(?:quantity|qty)\s*[:\-]?\s*(\d+(?:\.\d+)?)',
        r'hours\s*[:\-]\s*(\d+(?:\.\d+)?)',
    ]
    for pat in patterns:
        m = re.search(pat, text_lower, re.IGNORECASE)
        if m:
            try:
                val = float(m.group(1))
                if 0 < val <= 744:
                    return val
            except (ValueError, IndexError):
                pass
    return None


def process_with_document_intelligence(pdf_url: str) -> Dict:
    """
    Process PDF with Azure Document Intelligence
    
    Args:
        pdf_url: URL or path to PDF file
    
    Returns:
        Extracted data dictionary
    """
    import requests
    
    endpoint = os.environ.get('AZURE_DI_ENDPOINT')
    key = os.environ.get('AZURE_DI_KEY')
    
    if not endpoint or not key:
        raise ValueError("Document Intelligence not configured")
    
    # Use Document Intelligence REST API
    # This is a simplified version - adjust based on your needs
    analyze_url = f"{endpoint}formrecognizer/documentModels/prebuilt-invoice:analyze?api-version=2023-07-31"
    
    headers = {
        'Ocp-Apim-Subscription-Key': key,
        'Content-Type': 'application/json'
    }
    
    # Start analysis
    body = {
        'urlSource': pdf_url  # Or use base64Source for direct file upload
    }
    
    response = requests.post(analyze_url, headers=headers, json=body)
    response.raise_for_status()
    
    # Get operation ID
    operation_id = response.headers.get('Operation-Location').split('/')[-1]
    
    # Poll for results (simplified - implement proper polling)
    result_url = f"{endpoint}formrecognizer/documentModels/prebuilt-invoice/analyzeResults/{operation_id}?api-version=2023-07-31"
    
    # Wait and get results (implement retry logic)
    import time
    time.sleep(5)  # Simplified - implement proper polling
    
    result_response = requests.get(result_url, headers={'Ocp-Apim-Subscription-Key': key})
    result_response.raise_for_status()
    
    result = result_response.json()
    
    # Extract fields from result
    extracted_data = {}
    if 'analyzeResult' in result and 'documents' in result['analyzeResult']:
        doc = result['analyzeResult']['documents'][0]
        fields = doc.get('fields', {})
        
        extracted_data = {
            'invoice_number': fields.get('InvoiceId', {}).get('value'),
            'invoice_amount': fields.get('InvoiceTotal', {}).get('value'),
            'invoice_date': fields.get('InvoiceDate', {}).get('value'),
            'vendor_name': fields.get('VendorName', {}).get('value'),
            'vendor_address': fields.get('VendorAddress', {}).get('value'),
            # Add more fields as needed
        }
    
    return extracted_data

# ============================================================================
# iGentic Orchestrator Helpers
# ============================================================================

def process_with_igentic(extracted_data: Dict, invoice_id: str, session_id: Optional[str] = None) -> Dict:
    """
    Process invoice with iGentic Orchestrator (same payload as Backend igentic_json_post).
    
    Args:
        extracted_data: Dict with doc_name, full_text, extracted_text, timestamp, status
        invoice_id: Invoice ID (used as sessionId)
        session_id: Optional session ID for workflow continuation
    
    Returns:
        Orchestration result
    """
    import requests
    endpoint = os.environ.get('IGENTIC_ENDPOINT')
    if not endpoint:
        return {"status": "error", "error": "IGENTIC_ENDPOINT not configured"}
    payload = {
        "request": "Process invoice",
        "userInput": json.dumps(extracted_data),
        "sessionId": session_id or invoice_id,
    }
    try:
        response = requests.post(endpoint, json=payload, headers={"Content-Type": "application/json"}, timeout=120)
        response.raise_for_status()
        return response.json()
    except Exception as e:
        logger.exception("iGentic call failed")
        return {"status": "error", "error": str(e)}


def _extract_payment_details_from_igentic_response(resp: Dict) -> Optional[str]:
    """
    Extract payment details JSON from iGentic response (payment agent output).
    Agent instruction: "Display payment summary and give payment details in JSON format"
    Looks in result, display_text, agentResponses; handles responseData/orchestration_result wrappers.
    """
    def _search_in_text(raw: str) -> Optional[str]:
        if not raw or not isinstance(raw, str):
            return None
        raw_lower = raw.lower()
        if "payment" not in raw_lower and "ready for payment" not in raw_lower and "payment_summary" not in raw_lower:
            return None
        match = re.search(r'```json\s*(\{[\s\S]*?\})\s*```', raw, re.IGNORECASE | re.DOTALL)
        if match:
            try:
                parsed = json.loads(match.group(1))
                if isinstance(parsed, dict):
                    return json.dumps(parsed, indent=2)
            except (json.JSONDecodeError, ValueError):
                pass
        match = re.search(r'```\s*(\{[\s\S]*?"(?:payment|bank|account|amount|payee|payable|vendor|invoice)[\s\S]*?\})\s*```', raw, re.IGNORECASE | re.DOTALL)
        if match:
            try:
                parsed = json.loads(match.group(1))
                if isinstance(parsed, dict):
                    return json.dumps(parsed, indent=2)
            except (json.JSONDecodeError, ValueError):
                pass
        for m in re.finditer(r'```\s*(\{[\s\S]*?\})\s*```', raw, re.DOTALL):
            try:
                parsed = json.loads(m.group(1))
                if isinstance(parsed, dict) and any(
                    str(k).lower() in ("payment", "bank", "account", "amount", "payee", "payable", "vendor", "invoice", "summary")
                    for k in parsed.keys()
                ):
                    return json.dumps(parsed, indent=2)
            except (json.JSONDecodeError, ValueError):
                pass
        if "payment" in raw_lower:
            return raw[:3000] if len(raw) > 3000 else raw
        return None

    data = _get_igentic_searchable(resp)
    raw = data.get("result") or data.get("display_text") or data.get("displayText") or ""
    if isinstance(raw, dict):
        if any(
            str(k).lower() in ("payment", "payment_summary", "payment_details", "bank", "account", "payee")
            for k in raw.keys()
        ):
            return json.dumps(raw, indent=2)
        raw = json.dumps(raw)
    out = _search_in_text(raw)
    if out:
        return out
    agent_responses = data.get("agentResponses") or data.get("agent_responses")
    if isinstance(agent_responses, str):
        try:
            agent_responses = json.loads(agent_responses)
        except (json.JSONDecodeError, ValueError):
            pass
    if isinstance(agent_responses, list):
        for item in agent_responses:
            content = item.get("Content") or item.get("content") or item.get("result") or ""
            if isinstance(content, dict):
                content = json.dumps(content)
            out = _search_in_text(content)
            if out:
                return out
    raw = resp.get("result") or resp.get("display_text") or resp.get("displayText") or ""
    if isinstance(raw, dict):
        raw = json.dumps(raw)
    return _search_in_text(raw)


def validate_timesheet_hours_with_igentic(vendor_hours: float, timesheet: float, invoice_id: str) -> Optional[Dict]:
    """
    Send approved_hours (timesheet) and vendor_hours (invoice) to iGentic for comparison.
    Returns {"approval_status": "...", "hours_match": bool, "payment_details": str} or None on failure.
    When Complete/Ready for Payment, extracts payment details from response.
    """
    import requests
    endpoint = os.environ.get('IGENTIC_ENDPOINT')
    if not endpoint:
        return None
    user_input = {
        "vendor_hours": vendor_hours,
        "timesheet": timesheet,
        "invoice_id": invoice_id,
        "action": "compare_hours",
    }
    payload = {
        "request": "Validate timesheet hours",
        "userInput": json.dumps(user_input),
        "sessionId": invoice_id,
    }
    try:
        response = requests.post(endpoint, json=payload, headers={"Content-Type": "application/json"}, timeout=60)
        response.raise_for_status()
        result = response.json()
        data = result.get("responseData") or result.get("response_data") or result
        approval_status = data.get("approval_status") or (data.get("result") or {}).get("approval_status")
        hours_match = data.get("hours_match")
        if hours_match is None and isinstance(data.get("result"), dict):
            hours_match = data["result"].get("hours_match")
        if not approval_status:
            return None
        out = {"approval_status": approval_status, "hours_match": hours_match}
        if approval_status in ("Complete", "Ready for Payment", "ready for payment"):
            payment_details = _extract_payment_details_from_igentic_response(result)
            if payment_details:
                out["payment_details"] = payment_details
        return out
    except Exception as e:
        logger.warning("iGentic timesheet validation failed: %s", e)
    return None


def _compare_hours_locally(vendor_hours: float, timesheet: float) -> Dict:
    """
    Local fallback: compare timesheet vs vendor_hours per agent instructions.
    CASE 1: Match -> Complete; CASE 2: timesheet > invoice -> Need manual review;
    CASE 3: invoice > timesheet -> NEED APPROVAL.
    """
    try:
        v = float(vendor_hours) if vendor_hours is not None else 0
        t = float(timesheet) if timesheet is not None else 0
    except (TypeError, ValueError):
        return {"approval_status": "NEED APPROVAL", "hours_match": False}
    if abs(t - v) < 0.01:
        return {"approval_status": "Complete", "hours_match": True}
    if t > v:
        return {"approval_status": "Need manual review", "hours_match": False}
    return {"approval_status": "NEED APPROVAL", "hours_match": False}


# ============================================================================
# iGentic Response Parsing (CSV + JSON block)
# ============================================================================

def _get_igentic_searchable(resp: Dict) -> Dict:
    """
    Normalize iGentic response - unwrap responseData or orchestration_result wrappers.
    Returns the inner dict containing result, agentResponses, etc.
    """
    if not isinstance(resp, dict):
        return {}
    inner = resp.get("responseData") or resp.get("response_data")
    if not isinstance(inner, dict):
        orch = resp.get("orchestration_result") or resp.get("orchestrationResult")
        if isinstance(orch, dict):
            inner = orch.get("responseData") or orch.get("response_data") or orch
    if isinstance(inner, dict):
        return inner
    return resp


def extract_csv_from_igentic_response(orchestration_response: Dict) -> Optional[str]:
    """
    Extract CSV-style string from iGentic response.
    Looks in result, agentResponses, or display_text for CSV data.
    Handles responseData wrapper (iGentic API format).
    """
    import re
    data = _get_igentic_searchable(orchestration_response)
    
    # Check result field
    result = data.get("result") or data.get("display_text") or data.get("displayText") or ""
    if isinstance(result, dict):
        result = json.dumps(result)
    
    # Look for CSV pattern (header row with Invoice_Number, Vendor_Name, etc.)
    # Pattern: Invoice_Number followed by Vendor_Name (with optional comma/colon and whitespace)
    csv_pattern = r'Invoice_Number[,:]?\s*Vendor_Name[^\n]*\n[^\n]+(?:\n[^\n]+)*'
    match = re.search(csv_pattern, result, re.DOTALL | re.IGNORECASE)
    if match:
        csv_data = match.group(0)
        # Extract just the data rows (skip markdown code blocks if present)
        csv_data = re.sub(r'```[^\n]*\n', '', csv_data)
        csv_data = re.sub(r'```', '', csv_data)
        logger.info(f"Found CSV in result field: {csv_data[:200]}")
        return csv_data.strip()
    
    # Check agentResponses
    agent_responses = data.get("agentResponses") or data.get("agent_responses")
    if isinstance(agent_responses, str):
        try:
            agent_responses = json.loads(agent_responses)
        except Exception:
            pass
    
    if isinstance(agent_responses, list):
        for item in agent_responses:
            content = item.get("Content") or item.get("content") or ""
            if isinstance(content, str):
                match = re.search(csv_pattern, content, re.DOTALL | re.IGNORECASE)
                if match:
                    csv_data = match.group(0)
                    csv_data = re.sub(r'```[^\n]*\n', '', csv_data)
                    csv_data = re.sub(r'```', '', csv_data)
                    logger.info(f"Found CSV in agentResponses: {csv_data[:200]}")
                    return csv_data.strip()
    
    logger.warning(f"No CSV pattern found in iGentic response. Result preview: {str(result)[:500]}")
    return None


def parse_csv_to_dict(csv_string: str) -> Dict:
    """
    Parse CSV string from iGentic to structured dictionary.
    Expected CSV format:
    Invoice_Number,Vendor_Name,Resource_Name,Start_Date,End_Date,Invoice_Hours,Hourly_Rate,Total_Amount,Payment_Terms,Invoice_Date,Business_Unit,Project_Name
    20876,Sigmago Solutions Inc,Muneer - UI/UX designer,, ,144,169,24336.00,Net 30,2025-12-20,,Consulting services Invoice for Rishan
    """
    import csv
    import io
    
    if not csv_string:
        return {}
    
    # Clean up CSV string (remove markdown, extra whitespace)
    csv_string = csv_string.strip()
    csv_string = re.sub(r'```[^\n]*\n', '', csv_string)
    csv_string = re.sub(r'```', '', csv_string)
    
    lines = csv_string.split('\n')
    if len(lines) < 2:
        return {}
    
    # Parse header and first data row
    reader = csv.DictReader(io.StringIO(csv_string))
    try:
        row = next(reader)
    except StopIteration:
        return {}
    
    # Map CSV columns to our database fields (Vendor_Hours/Invoice_Hours -> invoice_hours)
    field_map = {
        'Invoice_Number': 'invoice_number',
        'Vendor_Name': 'vendor_name',
        'Resource_Name': 'resource_name',
        'Start_Date': 'start_date',
        'End_Date': 'end_date',
        'Invoice_Hours': 'invoice_hours',
        'Vendor_Hours': 'invoice_hours',
        'Vendor Hours': 'invoice_hours',
        'Hours': 'invoice_hours',
        'Total_Hours': 'invoice_hours',
        'Billable_Hours': 'invoice_hours',
        'Quantity': 'invoice_hours',
        'Hourly_Rate': 'hourly_rate',
        'Total_Amount': 'invoice_amount',
        'Payment_Terms': 'payment_terms',
        'Invoice_Date': 'invoice_date',
        'Business_Unit': 'business_unit',
        'Project_Name': 'project_name',
    }
    
    out = {}
    for csv_key, db_key in field_map.items():
        val = row.get(csv_key, '').strip()
        if val and val.lower() not in ('null', 'none', ''):
            # Convert numeric fields
            if db_key in ('invoice_hours', 'hourly_rate', 'invoice_amount'):
                try:
                    val = float(val.replace(',', ''))
                except (ValueError, AttributeError):
                    pass
            out[db_key] = val
    
    return out


def _parse_markdown_extracted_info(text: str) -> Dict:
    """Parse markdown-style - **Key:** value lines into a dict."""
    out = {}
    for m in re.finditer(r'[-*]\s*\*\*([^:*]+)\*\*\s*:\s*(.+?)(?=\n|$)', text):
        key = m.group(1).strip().replace(" ", "_")
        val = m.group(2).strip().split("(")[0].strip()
        if val.lower() in ("null", "none", ""):
            continue
        try:
            if "." in val and re.match(r"^\d+\.\d+$", val):
                val = float(val)
            elif val.isdigit():
                val = int(val) if int(val) == float(val) else float(val)
        except (ValueError, TypeError):
            pass
        out[key] = val
    return out


def extract_json_block_from_igentic_response(orchestration_response: Dict) -> Dict:
    """
    Extract structured fields from iGentic JSON block in result.
    iGentic often returns: **Structured JSON Output:** ```json { "Invoice_Number": "...", ... } ```
    Maps to our DB fields: invoice_number, vendor_name, invoice_amount, etc.
    """
    data = _get_igentic_searchable(orchestration_response)
    result = data.get("result") or data.get("display_text") or data.get("displayText") or ""
    if isinstance(result, dict):
        result = json.dumps(result)
    if not result:
        return {}

    # Extract ```json ... ``` block
    match = re.search(r'```json\s*(\{[\s\S]*?\})\s*```', result, re.IGNORECASE | re.DOTALL)
    if not match:
        match = re.search(r'```\s*(\{[\s\S]*?"Invoice_Number"[\s\S]*?\})\s*```', result, re.IGNORECASE | re.DOTALL)
    parsed = {}
    if match:
        try:
            parsed = json.loads(match.group(1))
        except (json.JSONDecodeError, ValueError):
            pass
    # Fallback: parse markdown bullets "- **Invoice_Hours:** 152"
    if not parsed and "Invoice_Hours" in result:
        parsed = _parse_markdown_extracted_info(result)

    field_map = {
        'Invoice_Number': 'invoice_number',
        'Vendor_Name': 'vendor_name',
        'Resource_Name': 'resource_name',
        'Start_Date': 'start_date',
        'End_Date': 'end_date',
        'Invoice_Hours': 'invoice_hours',
        'Vendor_Hours': 'invoice_hours',
        'VendorHours': 'invoice_hours',
        'Hours': 'invoice_hours',
        'Total_Hours': 'invoice_hours',
        'Billable_Hours': 'invoice_hours',
        'Total Hours': 'invoice_hours',
        'Quantity': 'invoice_hours',
        'Hourly_Rate': 'hourly_rate',
        'Total_Amount': 'invoice_amount',
        'Payment_Terms': 'payment_terms',
        'Invoice_Date': 'invoice_date',
        'Business_Unit': 'business_unit',
        'Project_Name': 'project_name',
    }
    out = {}
    for src_key, db_key in field_map.items():
        val = parsed.get(src_key)
        if val is None:
            continue
        if isinstance(val, str) and val.strip().lower() in ('null', 'none', ''):
            continue
        if db_key in ('invoice_hours', 'hourly_rate', 'invoice_amount') and val is not None:
            try:
                val = float(val) if not isinstance(val, (int, float)) else float(val)
            except (ValueError, TypeError):
                pass
        out[db_key] = val
    if out:
        logger.info("Extracted %d fields from iGentic JSON block: %s", len(out), list(out.keys()))
    return out


# Agent snake_case to DB column mapping (iGentic direct/flat output)
_IGENTIC_TO_DB = {
    "invoice_number": "invoice_number",
    "consultancy_name": "vendor_name",
    "resource_name": "resource_name",
    "pay_period_start": "start_date",
    "pay_period_end": "end_date",
    "vendor_hours": "invoice_hours",
    "approved_hours": "approved_hours",
    "pay_rate": "hourly_rate",
    "invoice_amount": "invoice_amount",
    "net_terms": "payment_terms",
    "invoice_date": "invoice_date",
    "due_date": "due_date",
    "business_unit": "business_unit",
    "project_name": "project_name",
    "template": "template",
    "approval_status": "approval_status",
    "status": "status",
}


def _extract_direct_igentic_fields(obj: Dict) -> Dict:
    """Extract from direct/flat iGentic object (snake_case agent output)."""
    out = {}
    for src_key, db_key in _IGENTIC_TO_DB.items():
        val = obj.get(src_key)
        if val is None:
            continue
        if isinstance(val, str) and val.strip().lower() in ("null", "none", ""):
            continue
        if db_key in ("invoice_hours", "hourly_rate", "invoice_amount", "approved_hours") and val is not None:
            try:
                val = float(val) if not isinstance(val, (int, float)) else float(val)
            except (ValueError, TypeError):
                pass
        out[db_key] = val
    return out


def extract_fields_from_igentic(orchestration_response: Dict) -> Dict:
    """
    Extract all structured fields from iGentic response for SQL/Excel update.
    Tries: 1) Direct flat object, 2) CSV, 3) JSON block in result. Merges status.
    """
    out = {}
    # 0) Direct flat object (iGentic agent snake_case output)
    for candidate in [orchestration_response, _get_igentic_searchable(orchestration_response)]:
        if isinstance(candidate, dict) and candidate.get("invoice_number"):
            direct = _extract_direct_igentic_fields(candidate)
            if direct:
                out.update(direct)
                logger.info("Extracted %d fields from iGentic direct object", len(direct))
                break
        # Also check responseData.result if it's a dict
        result = candidate.get("result") if isinstance(candidate, dict) else None
        if isinstance(result, dict) and result.get("invoice_number"):
            direct = _extract_direct_igentic_fields(result)
            if direct:
                out.update(direct)
                logger.info("Extracted %d fields from iGentic result dict", len(direct))
                break
    # 1) Try CSV
    if not out.get("invoice_number"):
        try:
            csv_string = extract_csv_from_igentic_response(orchestration_response)
            if csv_string:
                out.update(parse_csv_to_dict(csv_string))
        except Exception as e:
            logger.warning("CSV extraction failed: %s", e)
    # 2) If still no invoice_number, try JSON block (markdown+JSON format)
    if not out.get("invoice_number"):
        json_fields = extract_json_block_from_igentic_response(orchestration_response)
        out.update(json_fields)
    # 3) Status from result text
    data = _get_igentic_searchable(orchestration_response)
    raw = data.get("result") or data.get("display_text") or data.get("displayText") or ""
    if isinstance(raw, dict):
        raw = json.dumps(raw)
    text = (raw or "").lower()
    if "complete" in text or "ready for payment" in text:
        out["approval_status"] = "Complete"
        out["status"] = "Complete"
    elif "need approval" in text or "manual review" in text:
        out["approval_status"] = "NEED APPROVAL"
        out["status"] = "NEED APPROVAL"
    else:
        out.setdefault("approval_status", "Pending")
        out.setdefault("status", "Pending")
    return out


# ============================================================================
# Excel Helpers
# ============================================================================

def update_excel_file(invoice_id: str, invoice_data: Dict) -> None:
    """
    Update SharePoint Excel file with invoice data from CSV (iGentic output).
    Maps CSV fields: Invoice_Number, Vendor_Name, Invoice_Hours, Hourly_Rate, Total_Amount, etc.
    """
    from openpyxl import load_workbook
    import io
    
    # Download Excel from SharePoint
    excel_path = os.environ.get("SHAREPOINT_EXCEL_PATH") or "Invoices/Invoice_Register_Master.xlsx"
    server_relative_excel_url = _normalize_server_relative_url(excel_path)
    try:
        excel_content = download_file_from_sharepoint(server_relative_excel_url)
    except Exception as e:
        # Bubble up to caller so we don't log false success.
        raise FileNotFoundError(f"Excel file not found at {server_relative_excel_url}") from e
    
    # Load workbook
    wb = load_workbook(io.BytesIO(excel_content))
    ws = wb.active
    
    # Map invoice_data fields to Excel columns (adjust column indices based on your Excel structure)
    # Common columns: invoice_id, vendor_name, invoice_number, invoice_amount, invoice_hours, hourly_rate, status, invoice_date, etc.
    
    def _v(k, *alt):
        for a in [k] + list(alt):
            v = invoice_data.get(a)
            if v is not None and v != "":
                return v
        return None

    # Find row with invoice_id or append new row (add/update only, same file)
    row_found = False
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        if row[0].value == invoice_id:
            ws.cell(row=row_idx, column=2, value=_v('vendor_name', 'consultancy_name', 'vendor_id'))
            ws.cell(row=row_idx, column=3, value=_v('invoice_number'))
            ws.cell(row=row_idx, column=4, value=_v('invoice_amount'))
            ws.cell(row=row_idx, column=5, value=_v('status', 'approval_status'))
            ws.cell(row=row_idx, column=6, value=_v('invoice_hours', 'vendor_hours', 'approved_hours'))
            ws.cell(row=row_idx, column=7, value=_v('hourly_rate', 'pay_rate'))
            ws.cell(row=row_idx, column=8, value=_v('invoice_date'))
            ws.cell(row=row_idx, column=9, value=_v('resource_name'))
            ws.cell(row=row_idx, column=10, value=_v('project_name'))
            ws.cell(row=row_idx, column=11, value=_v('payment_terms', 'net_terms'))
            ws.cell(row=row_idx, column=12, value=_v('start_date', 'pay_period_start'))
            ws.cell(row=row_idx, column=13, value=_v('end_date', 'pay_period_end'))
            ws.cell(row=row_idx, column=14, value=_v('doc_name'))
            ws.cell(row=row_idx, column=15, value=_v('due_date'))
            ws.cell(row=row_idx, column=16, value=_v('notes', 'current_comments'))
            ws.cell(row=row_idx, column=17, value=_v('addl_comments'))
            row_found = True
            break

    if not row_found:
        ws.append([
            invoice_id,
            _v('vendor_name', 'consultancy_name', 'vendor_id'),
            _v('invoice_number'),
            _v('invoice_amount'),
            _v('status', 'approval_status'),
            _v('invoice_hours', 'vendor_hours', 'approved_hours'),
            _v('hourly_rate', 'pay_rate'),
            _v('invoice_date'),
            _v('resource_name'),
            _v('project_name'),
            _v('payment_terms', 'net_terms'),
            _v('start_date', 'pay_period_start'),
            _v('end_date', 'pay_period_end'),
            _v('doc_name'),
            _v('due_date'),
            _v('pdf_url'),
            _v('notes', 'current_comments'),
            _v('addl_comments'),
            datetime.utcnow().isoformat() if not invoice_data.get('created_at') else invoice_data.get('created_at'),
            datetime.utcnow().isoformat(),
        ])
    
    # Save and upload back to SharePoint
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    ctx = get_sharepoint_context()
    file = ctx.web.get_file_by_server_relative_url(server_relative_excel_url)
    file.save(output.read()).execute_query()
    
    logger.info(f"Updated Excel file with invoice {invoice_id} data")

# ============================================================================
# Logging Helpers
# ============================================================================

def save_complete_log(invoice_id: str, extracted_data: Dict, orchestration_result: Dict, event_type: str = 'upload') -> None:
    """Save complete audit trail (Document Intelligence + iGentic JSON) to SharePoint JSON_Logs for backup."""
    try:
        sql_record = None
        if os.environ.get('SQL_CONNECTION_STRING'):
            try:
                sql_record = get_invoice(invoice_id)
            except Exception:
                pass

        log_data = {
            'invoice_id': invoice_id,
            'timestamp': datetime.utcnow().isoformat(),
            'event_type': event_type,
            'extracted_data': extracted_data,
            'orchestration_result': orchestration_result,
            'database_record': sql_record
        }

        file_name = f'invoice_{invoice_id}_{event_type}.json'
        save_json_to_sharepoint(log_data, file_name)

        logger.info(f"Saved JSON log for invoice {invoice_id}")
    except Exception as e:
        logger.error(f"Failed to save JSON log: {str(e)}")

def save_status_change_log(invoice_id: str, old_status: str, new_status: str, changed_by: str) -> None:
    """Save status change log"""
    log_data = {
        'invoice_id': invoice_id,
        'timestamp': datetime.utcnow().isoformat(),
        'event_type': 'status_change',
        'old_status': old_status,
        'new_status': new_status,
        'changed_by': changed_by,
        'database_record': get_invoice(invoice_id)
    }
    
    file_name = f'invoice_{invoice_id}_status_change.json'
    save_json_to_sharepoint(log_data, file_name)
