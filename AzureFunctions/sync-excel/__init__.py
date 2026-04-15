import logging
import json
import os
import re
import threading
import io
import unicodedata
from datetime import datetime

import azure.functions as func
import openpyxl
import psycopg2
import psycopg2.extras

logger = logging.getLogger(__name__)

# Lazy-load so import errors surface in main() (logged + JSON 500) instead of failing at module load.
_helpers_mod = None
_helpers_lock = threading.Lock()


def _helpers():
    global _helpers_mod
    if _helpers_mod is not None:
        return _helpers_mod
    with _helpers_lock:
        if _helpers_mod is None:
            import shared.helpers as h

            _helpers_mod = h
    return _helpers_mod

# ── Constants ────────────────────────────────────────────────────────────────

MIN_TOKEN_LEN = 2          # skip single-char initials like "J."
SUFFIX_RE  = re.compile(
    r'\b(jr\.?|sr\.?|ii|iii|iv|ph\.?d\.?|md|esq\.?|cpa)\b', re.I)
PREFIX_RE  = re.compile(
    r'\b(mr\.?|mrs\.?|ms\.?|dr\.?|prof\.?)\b', re.I)

# Columns expected in the Excel file (case-insensitive match performed at runtime)
COL_FIRST      = 'first name'
COL_LAST       = 'last name'
COL_APPROVAL   = 'approval status'
COL_HOURS      = 'hour(s)'          # fallback: 'hours'
COL_FROM       = 'from time'        # used to derive pay period month
COL_DATE       = 'date'             # alternative pay-period column
COL_DIVISION   = 'division'
COL_CLIENT     = 'client name'
COL_PROJECT    = 'project name'

# ── Entry point ───────────────────────────────────────────────────────────────

def main(req: func.HttpRequest) -> func.HttpResponse:
    """Host must always get an HttpResponse; log any uncaught error (avoids opaque 'Failed' with no traceback)."""
    try:
        return _sync_excel_handler(req)
    except Exception as e:
        logger.exception("sync-excel unhandled error: %s", e)
        return func.HttpResponse(
            json.dumps({"error": str(e)}),
            status_code=500,
            mimetype="application/json",
            headers={"Access-Control-Allow-Origin": "*"},
        )


def _sync_excel_handler(req: func.HttpRequest) -> func.HttpResponse:
    logger.info("sync-excel triggered")

    # 1. Get uploaded file
    try:
        file_bytes = _get_file_bytes(req)
    except ValueError as e:
        return _err(400, str(e))

    try:
        q = getattr(req, "params", None)
        if q is not None and hasattr(q, "get"):
            filename = q.get("filename") or "timesheet.xlsx"
        else:
            filename = "timesheet.xlsx"
    except Exception:
        filename = "timesheet.xlsx"

    # ── 2. Parse Excel first (need row dates to name SharePoint file per month) ─
    try:
        rows = _parse_excel(file_bytes)
    except Exception as e:
        #logger.error("Excel parse error: %s", e)
        return _err(400, f"Could not parse Excel file: {e}")

    if not rows:
        return _err(400, "Excel file contained no data rows.")

    try:
        sharepoint_filename = _sharepoint_timesheet_filename(rows, filename)
    except ValueError as e:
        return _err(400, str(e))

    # 3. Save to SharePoint in background (fixed name per year-month replaces prior upload)
    sp_thread = threading.Thread(
        target=_save_to_sharepoint,
        args=(file_bytes, sharepoint_filename),
        daemon=True
    )
    sp_thread.start()

    # 4. Load Pending invoices from DB
    conn = None
    cursor = None
    try:
        conn = _helpers().get_sql_connection(
            cursor_factory=psycopg2.extras.RealDictCursor
        )
        cursor = conn.cursor()
        cursor.execute("""
            SELECT invoice_id, resource_name, vendor_name, employee_id, start_date, end_date,
                invoice_hours, approval_status, division, client_name,
                project_name_excel, payment_details
            FROM   invoices
            WHERE  LOWER(approval_status) = 'pending'
            AND  invoice_hours IS NOT NULL
        """)
        invoices = cursor.fetchall()

        # 5. Group Excel rows by person + month
        groups = _group_rows(rows)

        # 6. Match & update
        results = []
        for group_key, group in groups.items():
            _upsert_timesheet_hours_cache(group_key, group, cursor, conn)
            result = _process_group(group_key, group, invoices, cursor, conn)
            results.append(result)

    except Exception as e:
        logger.exception("sync-excel DB/processing error: %s", e)
        return _err(500, f"Database error: {e}")
    finally:
        # Always run on its own connection so it executes even if the main cursor path failed.
        try:
            _run_standalone_employee_id_hours_auto_approve()
        except Exception as e:
            logger.warning("Standalone auto-approval backfill skipped: %s", e)
        if cursor is not None:
            try:
                cursor.close()
            except Exception:
                pass
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass

    # 7. Return summary
    # Always refresh the monthly sync report (single canonical file per timesheet month).
    sp_thread.join(timeout=2)

    unmatched_ambiguous = [r for r in results if r['status'] in ('UNMATCHED', 'AMBIGUOUS', 'PERIOD_MISMATCH')]
    report_thread = threading.Thread(
        target=_upload_comparison_report,
        args=(unmatched_ambiguous, results, invoices, sharepoint_filename, groups),
        daemon=True
    )
    report_thread.start()
    
    raw_sheet_hours = sum(
    sum(_to_float((_get_col(r, COL_HOURS) or _get_col(r, 'hours') or '0').strip()) for r in group)
    for group in groups.values()
    )

    summary = {
            "processed":    len(results),
            "matched":      sum(1 for r in results if r['status'] == 'MATCHED'),
            "need_approval":sum(1 for r in results if r['status'] == 'Need Approval'),
            "pending":      sum(1 for r in results if r['status'] == 'PENDING'),
            "unmatched":    sum(1 for r in results if r['status'] == 'UNMATCHED'),
            "ambiguous":    sum(1 for r in results if r['status'] == 'AMBIGUOUS'),
            "skipped_not_pending": sum(1 for r in results if r['status'] == 'SKIPPED_NOT_PENDING'),
            "period_mismatch": sum(1 for r in results if r['status'] == 'PERIOD_MISMATCH'),
            # Hour tallies
            "total_hours": sum(r.get('total_hours', 0) for r in results),
            "matched_hours": sum(r.get('total_hours', 0) for r in results if r['status'] == 'MATCHED'),
            "need_approval_hours": sum(r.get('total_hours', 0) for r in results if r['status'] == 'Need Approval'),
            "pending_hours": sum(r.get('total_hours', 0) for r in results if r['status'] == 'PENDING'),
            "unmatched_hours": sum(r.get('total_hours', 0) for r in results if r['status'] == 'UNMATCHED'),
            "ambiguous_hours": sum(r.get('total_hours', 0) for r in results if r['status'] == 'AMBIGUOUS'),
            "period_mismatch_hours": sum(r.get('total_hours', 0) for r in results if r['status'] == 'PERIOD_MISMATCH'),
            "details":      results
}

    return func.HttpResponse(
        json.dumps(summary, default=str),
        status_code=200,
        mimetype="application/json",
        headers={"Access-Control-Allow-Origin": "*"}
    )


# File extraction

def _get_file_bytes(req: func.HttpRequest) -> bytes:

    content_type = req.headers.get('Content-Type', '')

    if 'multipart/form-data' in content_type:
        files = req.files
        if not files:
            raise ValueError("No file found in multipart request.")
        key  = next(iter(files))
        data = files[key].read()
        if not data:
            raise ValueError("Uploaded file is empty.")
        return data

    # Raw binary body
    data = req.get_body()
    if not data:
        raise ValueError("Request body is empty. Send the Excel file as binary body "
                         "or as multipart/form-data.")
    return data


# Excel parsing 

def _parse_excel(file_bytes: bytes) -> list:

    wb  = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws  = wb.active

    headers = []
    rows    = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip().lower() if c is not None else '' for c in row]
            continue

        if all(c is None for c in row):
            continue    # skip blank rows

        record = dict(zip(headers, row))
        rows.append(record)

    wb.close()
    return rows


# Grouping 

def _group_rows(rows: list) -> dict:

    groups = {}

    for row in rows:
        first = _get_col(row, COL_FIRST)
        last  = _get_col(row, COL_LAST)
        if not first and not last:
            continue

        yr, mo = _extract_month_year(row)
        if yr is None:
            yr, mo = 0, 0   # group together rows with no date

        emp = _row_employee_id(row)
        # Group key: column E employee id + month (standard template). If E is empty, name + month.
        if emp:
            key = ('emp', emp.lower().strip(), yr, mo)
        else:
            key = ('name', _normalise(first), _normalise(last), yr, mo)
        groups.setdefault(key, []).append(row)

    return groups


def _extract_month_year(row: dict):

    for col in (COL_DATE,'date'):
        val = _get_col(row, col)
        if val:
            dt = _parse_date(val)
            if dt:
                return dt.year, dt.month
    return None, None


def _unambiguous_year_month_from_rows(rows: list) -> tuple:
    """
    SharePoint canonical name only when every dated row agrees on one (year, month).
    Avoids naming a file from a 'dominant' month when the workbook mixes periods.
    """
    seen = set()
    for row in rows:
        yr, mo = _extract_month_year(row)
        if yr and mo:
            seen.add((yr, mo))
    if len(seen) == 1:
        return next(iter(seen))
    if len(seen) == 0:
        raise ValueError(
            "Timesheet must contain date values for one month (used to replace that month's existing file)."
        )
    raise ValueError(
        f"Timesheet spans multiple months {sorted(seen)}. Upload one month per file so it can replace that month's existing timesheet."
    )


def _sharepoint_timesheet_filename(rows: list, original_filename: str) -> str:
    """
    Stable name timesheet_YYYY_MM.ext from a single month workbook.
    Graph PUT to the same path overwrites the previous timesheet for that month.
    """
    yr, mo = _unambiguous_year_month_from_rows(rows)
    ext = "xlsx"
    if "." in (original_filename or ""):
        ext = original_filename.rsplit(".", 1)[-1].lower()
    if ext not in ("xlsx", "xls", "xlsm"):
        ext = "xlsx"
    return f"timesheet_{yr}_{mo:02d}.{ext}"


def _parse_date(val):
    if isinstance(val, datetime):
        return val
    if hasattr(val, 'year'):        # date object
        return val
    s = str(val).strip()
    for fmt in ('%d/%m/%Y','%d-%m-%Y','%Y-%m-%d','%Y/%m/%d','%m/%d/%Y','%d/%m/%Y',):
        try:
            return datetime.strptime(s[:10], fmt)
        except ValueError:
            pass
    return None


# Name normalisation & regex matching

def _fold_unicode_accents(s: str) -> str:
    """João/joao/JOAO → same letters for matching (human-like)."""
    if not s:
        return ""
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def _normalise(name: str) -> str:
    s = _fold_unicode_accents(name or "")
    s = s.lower()
    s = PREFIX_RE.sub('', s)
    s = SUFFIX_RE.sub('', s)
    s = re.sub(r"[-'\u2018\u2019\u201a\u201b`]", ' ', s)
    s = re.sub(r'[^a-z\s]', '', s)
    return re.sub(r'\s+', ' ', s).strip()


def _tokenise(name: str) -> list:
    return [t for t in _normalise(name).split() if len(t) >= MIN_TOKEN_LEN]


def _any_token_in(tokens: list, normed_db: str) -> bool:
    return any(re.search(r'\b' + re.escape(t) + r'\b', normed_db) for t in tokens)

# Pay period check 

def _pay_period_matches(invoice, year: int, month: int) -> bool:

    #if year == 0:
    #   return True     # no date in timesheet — don't block on period

    for field in ('start_date', 'end_date'):
        val = invoice.get(field)
        if val:
            dt = _parse_date(str(val))
            if dt and dt.year == year and dt.month == month:
                return True
    return False


def _match_invoice_by_emp_id(emp_id: str, invoices: list):
    """Match only on invoice.employee_id (Excel column E). Never row/S.No. or other columns."""
    emp = (emp_id or '').strip().lower()
    if not emp:
        return None, 'UNMATCHED'
    matches = [inv for inv in invoices if str(inv.get('employee_id') or '').strip().lower() == emp]
    if len(matches) == 1:
        return matches[0], 'MATCHED'
    if len(matches) > 1:
        return None, 'AMBIGUOUS'
    return None, 'UNMATCHED'


def _name_tokens(name: str) -> list:
    return [t for t in _normalise(name).split() if len(t) > 1]


def _token_similar(a: str, b: str) -> bool:
    if a == b:
        return True
    la, lb = len(a), len(b)
    if la < 2 or lb < 2:
        return False
    # Short given names (e.g. Ana, Li) — only accept exact match after normalise
    if max(la, lb) < 4:
        return a == b
    from difflib import SequenceMatcher
    return SequenceMatcher(None, a, b).ratio() >= 0.84


def _human_name_gate_match(a_name: str, b_name: str) -> bool:
    """
    Human-like name equivalence:
    - at least 2 meaningful tokens match,
    - small spelling mistakes allowed,
    - single-letter initials ignored,
    - short partials like 'jee' do not pass.
    """
    a_t = _name_tokens(a_name)
    b_t = _name_tokens(b_name)
    if not a_t or not b_t:
        return False
    used = set()
    matches = 0
    for a in a_t:
        for i, b in enumerate(b_t):
            if i in used:
                continue
            if _token_similar(a, b):
                used.add(i)
                matches += 1
                break
    return matches >= 2


def _closest_name_invoice(first: str, last: str, invoices: list):
    """
    Fallback when column E did not match any invoice.employee_id: compare Excel first/last
    to invoice.resource_name only (human-name gate + similarity). Does not use S.No. or ids.
    """
    from difflib import SequenceMatcher
    q = _normalise(f"{first} {last}".strip())
    if not q:
        return None, 'UNMATCHED'
    scored = []
    for inv in invoices:
        db = _normalise(inv.get('resource_name') or '')
        if not db:
            continue
        if not _human_name_gate_match(q, db):
            continue
        score = SequenceMatcher(None, q, db).ratio()
        scored.append((score, inv))
    if not scored:
        return None, 'UNMATCHED'
    scored.sort(key=lambda x: x[0], reverse=True)
    best_score, best_inv = scored[0]
    second_score = scored[1][0] if len(scored) > 1 else 0.0
    # Conservative threshold + margin to avoid accidental wrong-person selection.
    if best_score < 0.72:
        return None, 'UNMATCHED'
    if second_score >= best_score - 0.03:
        return None, 'AMBIGUOUS'
    return best_inv, 'Need Approval'

def _match_invoice(first: str, last: str, invoices: list):
    """
    Returns (invoice, status) where status is one of:
    MATCHED       — both first and last name found in DB resource_name
    Need Approval — only first OR last name found
    AMBIGUOUS     — multiple invoices matched
    PERIOD_MISMATCH — name found in DB but under a different month
    UNMATCHED     — no match found
    """
    first_toks = _tokenise(first)
    last_toks  = _tokenise(last)

    # Pass 1: both first AND last token found in DB name
    full = [
        inv for inv in invoices
        if (db := _normalise(inv['resource_name'] or ''))
        and _any_token_in(first_toks, db)
        and _any_token_in(last_toks,  db)
    ]
    if len(full) == 1:
        return full[0], 'MATCHED'
    if len(full) > 1:
        return None, 'AMBIGUOUS'

    # Pass 2: partial match (one side only)
    partial = [
        inv for inv in invoices
        if (db := _normalise(inv['resource_name'] or ''))
        and (_any_token_in(first_toks, db) or _any_token_in(last_toks, db))
    ]
    if len(partial) == 1:
        return partial[0], 'Need Approval'
    if len(partial) > 1:
        return None, 'AMBIGUOUS'

    return None, 'UNMATCHED'



# Core processing 

def _process_group(group_key, group, invoices, cursor, conn) -> dict:
    first = _normalise(_get_col(group[0], COL_FIRST))
    last = _normalise(_get_col(group[0], COL_LAST))
    if group_key and group_key[0] == 'emp':
        _, _, yr, mo = group_key
    else:
        _, _, _, yr, mo = group_key
    # Timesheet employee id: always Excel column E (see _employee_id_from_column_e).
    emp_id = _employee_id_from_group(group)

    # Match to DB: (1) invoice.employee_id == column E, then month write guard.
    # (2) If no invoice has that id (or id path did not attach a row), closest name → month guard.
    month_invoices = [inv for inv in invoices if _pay_period_matches(inv, yr, mo)]
    inv = None
    match_status = 'UNMATCHED'
    if emp_id:
        inv_all, status_all = _match_invoice_by_emp_id(emp_id, invoices)
        if status_all == 'MATCHED':
            # Write guard: only same-month invoice can be updated
            same_month_emp = [
                i for i in month_invoices
                if str(i.get('employee_id') or '').strip().lower() == str(emp_id).strip().lower()
            ]
            if len(same_month_emp) == 1:
                inv, match_status = same_month_emp[0], 'MATCHED'
            elif len(same_month_emp) > 1:
                inv, match_status = None, 'AMBIGUOUS'
            else:
                inv, match_status = None, 'PERIOD_MISMATCH'
        elif status_all == 'AMBIGUOUS':
            inv, match_status = None, 'AMBIGUOUS'

    # Closest name only when column E did not match any pending invoice (or no emp_id).
    if inv is None and match_status in ('UNMATCHED',):
        missing_eid_all = [i for i in invoices if not str(i.get('employee_id') or '').strip()]
        candidates_all = missing_eid_all or invoices
        inv_all, status_all = _closest_name_invoice(first, last, candidates_all)
        if status_all == 'Need Approval' and inv_all is not None:
            same_month_name = [
                i for i in month_invoices
                if _human_name_gate_match(f"{first} {last}".strip(), i.get('resource_name') or '')
            ]
            if len(same_month_name) == 1:
                inv, match_status = same_month_name[0], 'Need Approval'
            elif len(same_month_name) > 1:
                inv, match_status = None, 'AMBIGUOUS'
            else:
                inv, match_status = None, 'PERIOD_MISMATCH'
        else:
            inv, match_status = None, status_all

    base = {
        'excel_name': f"{first} {last}",
        'excel_first': first,   
        'excel_last':  last,
        'year': yr,
        'month': mo,
        'row_count': len(group),
        'group_key': list(group_key),
        'employee_id': emp_id,
    }

    if match_status in ('UNMATCHED', 'AMBIGUOUS', 'PERIOD_MISMATCH'):
        total_hours = sum(_to_float((_get_col(r, COL_HOURS) or _get_col(r, 'hours') or '0').strip()) for r in group)
        return {**base, 'status': match_status, 'invoice_id': None, 'total_hours': total_hours}  # ADD total_hours


    # Guard: invoice must still be Pending
    current_status = (inv.get('approval_status') or '').strip().lower()
    if current_status != 'pending':
        total_hours = sum(_to_float((_get_col(r, COL_HOURS) or _get_col(r, 'hours') or '0').strip()) for r in group)
        return {**base,
                'status': 'SKIPPED_NOT_PENDING',
                'invoice_id': str(inv['invoice_id']),
                'db_status': inv.get('approval_status'),
                'total_hours': total_hours}  


    # Evaluate approval across all rows in this group
    approval_vals = [
        (_get_col(r, COL_APPROVAL) or '').strip().lower()
        for r in group
    ]
    all_approved = all(v == 'approved' for v in approval_vals)
    has_approved = any(v == 'approved' for v in approval_vals)
    has_non_approved = any(v != 'approved' for v in approval_vals)

    # Collect extra dimension fields from the group (take first non-empty value)
    division     = _first_val(group, COL_DIVISION)
    client_name  = _first_val(group, COL_CLIENT)
    project_name_excel = _first_val(group, COL_PROJECT)
    if all_approved and match_status == 'MATCHED':
        total_hours = sum(_to_float((_get_col(r, COL_HOURS) or _get_col(r, 'hours') or '0').strip()) for r in group)
        vendor_hrs = _to_float(str(inv.get('invoice_hours') or '0').strip())
        hours_match = abs(total_hours - vendor_hrs) < 0.5

        new_status  = 'Approved' if hours_match else 'Need Approval'
        payment_details = None

        if new_status == 'Approved':
            try:
                h = _helpers()
                ok_result = h.continue_igentic_session(
                    inv['invoice_id'],
                    "payment details",
                    request_label=f"The Approved hours is {total_hours}. Get payment details",
                )
                payment_details = h._extract_payment_details_from_igentic_response(
                    ok_result
                )
                if payment_details:
                    logger.info("Payment details extracted and saved from Excel Timesheet Update")
            except Exception as igentic_err:
                logger.warning("iGentic payment details call failed; continuing without payment details: %s", igentic_err)


        _invoice_update_with_employee_sync(cursor, conn, inv, {
            'approved_hours':  total_hours,
            'approval_status': new_status,
            'division':        division,
            'client_name':     client_name,
            'project_name_excel':    project_name_excel,
            'payment_details': payment_details,
        }, emp_id)



        return {**base,
                'status':         'MATCHED',
                'invoice_id':   str(inv['invoice_id']),
                'approved_hours': total_hours,
                'invoice_hours':   vendor_hrs,
                'new_db_status':  new_status,
                'matched_to':     inv.get('resource_name'),
                'db_employee_id': inv.get('employee_id')}
    
    # Case 2: Closest-name fallback path (employee id missing/unmatched)
    elif match_status == 'Need Approval':
        total_hours = sum(
            _to_float((_get_col(r, COL_HOURS) or _get_col(r, 'hours') or '0').strip())
            for r in group
        )
        logger.info("=== NEED APPROVAL (CLOSEST NAME) === Person: %s %s | Hours: %s | Invoice: %s",
                    first, last, total_hours, inv['invoice_id'])

        _invoice_update_with_employee_sync(cursor, conn, inv, {
            'approval_status':    'Need Approval',
            'approved_hours':     total_hours,
            'division':           division,
            'client_name':        client_name,
            'project_name_excel': project_name_excel,
        }, emp_id)
        return {**base,
                'status':         'Need Approval',
                'invoice_id':     str(inv['invoice_id']),
                'approved_hours': total_hours,
                'new_db_status':  'Need Approval',
                'matched_to':     inv.get('resource_name'),
                'db_employee_id': inv.get('employee_id')}
    
    #Case 3: Both names matched + mixed approval (some approved, some not)
    elif has_approved and has_non_approved:
        # Mixed → flag but still write dimension fields
        _invoice_update_with_employee_sync(cursor, conn, inv, {
            'approval_status': 'Need Approval',
            'division':        division,
            'client_name':     client_name,
            'project_name_excel':    project_name_excel,
        }, emp_id)
        return {**base,
                'status':        'Need Approval',
                'invoice_id':  str(inv['invoice_id']),
                'new_db_status': 'Need Approval',
                'matched_to':    inv.get('resource_name'),
                'db_employee_id': inv.get('employee_id')}

    else:
    # All still Pending — update dimensions only, leave approval_status
        total_hours = sum(_to_float((_get_col(r, COL_HOURS) or _get_col(r, 'hours') or '0').strip()) for r in group)
        _invoice_update_with_employee_sync(cursor, conn, inv, {
            'division':    division,
            'client_name': client_name,
            'project_name_excel':project_name_excel,
        }, emp_id)
        return {**base,
                'status':       'PENDING',
                'invoice_id': str(inv['invoice_id']),
                'matched_to':   inv.get('resource_name'),
                'db_employee_id': inv.get('employee_id'),
                'total_hours': total_hours} 



# DB helpers

def _invoice_update_with_employee_sync(cursor, conn, inv: dict, fields: dict, emp_id):
    """
    Apply invoice update. Timesheet employee_id is used only when the invoice row
    has no employee_id yet (SOW/upload path takes priority; SOW is not updated from timesheet).
    """
    existing_eid = (inv.get("employee_id") or "").strip()
    if emp_id and not existing_eid:
        fields = dict(fields)
        fields["employee_id"] = emp_id
        inv["employee_id"] = emp_id
    _write_update(cursor, conn, inv["invoice_id"], fields)


def _write_update(cursor, conn, invoice_id, fields: dict):
    """Build a dynamic UPDATE for only the provided fields."""
    # Remove None values
    fields = {k: v for k, v in fields.items() if v is not None}
    if not fields:
        return

    set_clause = ', '.join(f"{col} = %s" for col in fields)
    values     = list(fields.values()) + [invoice_id]

    cursor.execute(
        f"UPDATE invoices SET {set_clause}, excel_updated_at = NOW() WHERE invoice_id = %s",
        values
    )
    conn.commit()


# SharePoint upload

def _save_to_sharepoint(file_bytes: bytes, filename: str):
    """
    Upload the Excel file to SharePoint using the existing certificate-based
    context from shared/helpers.py.  Runs in a background thread so it does
    not block the HTTP response.  Errors are logged, not re-raised.
    """
    try:
        folder = os.environ.get('SP_FOLDER_PATH', 'Timesheet')
        url = _helpers().upload_excel_to_sharepoint(file_bytes, filename, folder)
        if url:
            logger.info("SharePoint upload OK: %s → %s", filename, url)
        else:
            logger.warning("SharePoint upload completed but returned no URL for: %s", filename)
    except Exception as e:
        logger.error("SharePoint upload error for '%s': %s", filename, e)


def _get_col(row: dict, col_name: str) -> str:
    """Case-insensitive column lookup."""
    key = col_name.lower()
    for k, v in row.items():
        if k.lower() == key and v is not None:
            return str(v).strip()
    return ''


def _first_val(group: list, col_name: str) -> str:
    """Return the first non-empty value for a column across a group of rows."""
    for row in group:
        val = _get_col(row, col_name)
        if val:
            return val
    return None


# Standard vendor layout: payroll / contractor id is read only from Excel column E (0-based index 4).
# DB matching uses that value against invoices.employee_id; if no row matches, closest name is used.
_EXCEL_EMPLOYEE_ID_COL_INDEX = 4


def _employee_id_from_column_e(row: dict) -> str:
    """Physical column E (5th column); not S.No. / line / serial columns elsewhere in the sheet."""
    v = _value_from_excel_col_index(row, _EXCEL_EMPLOYEE_ID_COL_INDEX)
    if v is None or not str(v).strip():
        return ""
    return str(v).strip()


def _employee_id_from_group(group: list):
    """First non-empty column E in the group (same person, same id on each row)."""
    if not group:
        return None
    for row in group:
        eid = _employee_id_from_column_e(row)
        if eid:
            return eid
    return None


def _row_employee_id(row: dict):
    eid = _employee_id_from_column_e(row)
    return eid if eid else None


def _value_from_excel_col_index(row: dict, idx0: int):
    """Return value from zero-based Excel column index from parsed row dict order."""
    if row is None or idx0 < 0:
        return None
    try:
        items = list(row.items())
        if idx0 >= len(items):
            return None
        _, val = items[idx0]
        return str(val).strip() if val is not None else None
    except Exception:
        return None


def _approved_hours_from_group(group: list) -> float:
    total = 0.0
    for r in group:
        approval = (_get_col(r, COL_APPROVAL) or "").strip().lower()
        if approval == "approved":
            total += _to_float((_get_col(r, COL_HOURS) or _get_col(r, "hours") or "0").strip())
    return total


def _upsert_timesheet_hours_cache(group_key, group: list, cursor, conn) -> None:
    """
    Persist timesheet approved-hours aggregates by month so invoices uploaded later
    can backfill approved_hours (look-back behavior).
    """
    if not group:
        return
    approved_hours = _approved_hours_from_group(group)
    if approved_hours <= 0:
        return

    if group_key and group_key[0] == 'emp':
        _, emp_id, yr, mo = group_key
        resource_name = _normalise(f"{_get_col(group[0], COL_FIRST)} {_get_col(group[0], COL_LAST)}")
    else:
        _, first, last, yr, mo = group_key
        emp_id = _employee_id_from_group(group)
        resource_name = _normalise(f"{first} {last}")

    if not yr or not mo:
        return

    division = _first_val(group, COL_DIVISION)
    client_name = _first_val(group, COL_CLIENT)
    project_name_excel = _first_val(group, COL_PROJECT)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS timesheet_hours_cache (
            cache_id BIGSERIAL PRIMARY KEY,
            year INT NOT NULL,
            month INT NOT NULL,
            employee_id TEXT NULL,
            resource_name TEXT NULL,
            approved_hours NUMERIC NULL,
            division TEXT NULL,
            client_name TEXT NULL,
            project_name_excel TEXT NULL,
            updated_at TIMESTAMP DEFAULT NOW()
        )
    """)
    cursor.execute("ALTER TABLE timesheet_hours_cache ADD COLUMN IF NOT EXISTS division TEXT NULL")
    cursor.execute("ALTER TABLE timesheet_hours_cache ADD COLUMN IF NOT EXISTS client_name TEXT NULL")
    cursor.execute("ALTER TABLE timesheet_hours_cache ADD COLUMN IF NOT EXISTS project_name_excel TEXT NULL")

    if emp_id:
        cursor.execute("""
            DELETE FROM timesheet_hours_cache
            WHERE year = %s AND month = %s
              AND LOWER(TRIM(COALESCE(employee_id, ''))) = LOWER(TRIM(%s))
        """, (yr, mo, str(emp_id).strip()))
    else:
        cursor.execute("""
            DELETE FROM timesheet_hours_cache
            WHERE year = %s AND month = %s
              AND employee_id IS NULL
              AND LOWER(TRIM(COALESCE(resource_name, ''))) = LOWER(TRIM(%s))
        """, (yr, mo, resource_name))

    cursor.execute("""
        INSERT INTO timesheet_hours_cache (
            year, month, employee_id, resource_name, approved_hours,
            division, client_name, project_name_excel, updated_at
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW())
    """, (
        yr, mo, (str(emp_id).strip() if emp_id else None), resource_name, approved_hours,
        division, client_name, project_name_excel
    ))
    conn.commit()


def _err(code: int, msg: str) -> func.HttpResponse:
    return func.HttpResponse(
        json.dumps({'error': msg}),
        status_code=code,
        mimetype='application/json',
        headers={"Access-Control-Allow-Origin": "*"}
    )

# Comparison report

# Comparison report

def _generate_comparison_report(unmatched_results: list, all_results: list, db_invoices: list, source_filename: str, groups: dict) -> bytes:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    hdr_font  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    thin      = Side(style='thin', color='CCCCCC')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    center    = Alignment(horizontal='center', vertical='center')
    wrap      = Alignment(wrap_text=True, vertical='top')

    def fill(hex_color):
        return PatternFill('solid', start_color=hex_color)

    def write_header(ws, row_num, labels, hex_color):
        for c, label in enumerate(labels, 1):
            cell = ws.cell(row=row_num, column=c, value=label)
            cell.font, cell.fill, cell.alignment, cell.border = hdr_font, fill(hex_color), center, border

    def style_cell(cell, hex_color=None):
        cell.alignment, cell.border = wrap, border
        if hex_color:
            cell.fill = fill(hex_color)

    # Calculate total hours for each category
    matched_results = [r for r in all_results if r['status'] == 'MATCHED']
    
    total_hours_matched = sum(_to_float(r.get('approved_hours', 0)) for r in matched_results)
    
    # Calculate unmatched hours from the groups
    total_hours_unmatched = 0
    for result in unmatched_results:
        key = tuple(result.get('group_key') or (
            result.get('excel_first', ''), result.get('excel_last', ''),
            result.get('year', 0), result.get('month', 0)
        ))
        group = groups.get(key, [])
        for row in group:
            hours = _to_float((_get_col(row, COL_HOURS) or _get_col(row, 'hours') or '0').strip())
            total_hours_unmatched += hours
    
    total_hours_pending = sum(_to_float(inv.get('invoice_hours', 0) or 0) for inv in db_invoices)

    # ── Sheet 1: Summary (Quality Check) ─────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = 'Summary'
    
    ws_summary.merge_cells('A1:C1')
    t_sum = ws_summary['A1']
    t_sum.value = f"Excel Sync Status - Quality Check Summary  |  {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    t_sum.font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    t_sum.fill = fill('34495E')
    t_sum.alignment = center
    ws_summary.row_dimensions[1].height = 25

    # Summary data
    summary_data = [
        ('Source File', source_filename, ''),
        ('Upload Time', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), ''),
        ('', '', ''),
        ('Category', 'Total Hours', 'Description'),
        ('Pending DB Invoices', f'{total_hours_pending:.2f}', 'All pending invoices in database'),
        ('Matched Data', f'{total_hours_matched:.2f}', 'Successfully matched and approved from Excel'),
        ('Unmatched Data', f'{total_hours_unmatched:.2f}', 'Excel entries with no DB match'),
        ('', '', ''),
        ('Total Matched Hours', f'{total_hours_matched:.2f}', 'Hours processed from timesheet'),
        ('Total Unmatched Hours', f'{total_hours_unmatched:.2f}', 'Hours not matched to DB'),
        ('Total Pending Hours', f'{total_hours_pending:.2f}', 'Hours awaiting approval in DB'),
    ]

    for r_idx, (label, value, desc) in enumerate(summary_data, start=2):
        if r_idx == 2 or r_idx == 9:  # Empty rows (adjusted for new upload time row)
            continue
        
        cell_a = ws_summary.cell(row=r_idx, column=1, value=label)
        cell_b = ws_summary.cell(row=r_idx, column=2, value=value)
        cell_c = ws_summary.cell(row=r_idx, column=3, value=desc)
        
        if r_idx == 5:  # Header row (adjusted for new row)
            cell_a.font = cell_b.font = cell_c.font = hdr_font
            cell_a.fill = cell_b.fill = cell_c.fill = fill('5DADE2')
            cell_a.alignment = cell_b.alignment = cell_c.alignment = center
        elif r_idx in [10, 11, 12]:  # Total rows (adjusted for new row)
            cell_a.font = Font(name='Arial', bold=True, size=11)
            cell_b.font = Font(name='Arial', bold=True, size=11)
            cell_a.fill = cell_b.fill = cell_c.fill = fill('D5F4E6')
        elif r_idx in [3, 4]:  # Source file and upload time rows
            cell_a.font = Font(name='Arial', bold=True, size=10)
            cell_b.font = Font(name='Arial', size=10)
        else:
            cell_a.alignment = cell_c.alignment = wrap
        
        cell_a.border = cell_b.border = cell_c.border = border

    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 18
    ws_summary.column_dimensions['C'].width = 45

    # ── Sheet 2: Pending DB Invoices ─────────────────────────────────────────
    ws2 = wb.create_sheet('Pending DB')
    ws2.freeze_panes = 'A3'

    ws2.merge_cells('A1:I1')
    t2 = ws2['A1']
    t2.value = f"All Pending Invoices in DB  |  {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    t2.font, t2.fill, t2.alignment = Font(name='Arial', bold=True, size=12, color='FFFFFF'), fill('2C3E50'), center
    ws2.row_dimensions[1].height = 22

    cols2 = ['Invoice ID', 'Resource Name', 'Employee ID', 'Start Date', 'End Date',
             'Invoice Hours', 'Approval Status', 'Division', 'Client Name']
    write_header(ws2, 2, cols2, '2980B9')

    for r_idx, inv in enumerate(db_invoices, start=3):
        row_data = [
            str(inv.get('invoice_id', '')),
            inv.get('resource_name', ''),
            inv.get('employee_id', ''),
            str(inv.get('start_date') or ''),
            str(inv.get('end_date') or ''),
            inv.get('invoice_hours', ''),
            inv.get('approval_status', ''),
            inv.get('division', ''),
            inv.get('client_name', ''),
        ]
        for c_idx, val in enumerate(row_data, 1):
            style_cell(ws2.cell(row=r_idx, column=c_idx, value=val))

    for col, width in zip('ABCDEFGHI', [14, 28, 14, 13, 13, 13, 16, 18, 24]):
        ws2.column_dimensions[get_column_letter(ord(col) - 64)].width = width

    # ── Sheet 3: Matched ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet('Matched')
    ws3.freeze_panes = 'A3'

    ws3.merge_cells('A1:K1')
    t3 = ws3['A1']
    t3.value = f"Matched Invoices  |  {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    t3.font, t3.fill, t3.alignment = Font(name='Arial', bold=True, size=12, color='FFFFFF'), fill('2C3E50'), center
    ws3.row_dimensions[1].height = 22

    cols3 = ['Excel Name', 'Timesheet Employee ID', 'Matched To (DB)', 'DB Employee ID', 'Invoice ID', 'Year', 'Month',
             'Approved Hours', 'Invoice Hours', 'New DB Status', 'Row Count']
    write_header(ws3, 2, cols3, '27AE60')

    for r_idx, result in enumerate(matched_results, start=3):
        row_data = [
            result.get('excel_name', ''),
            result.get('employee_id', ''),
            result.get('matched_to', ''),
            result.get('db_employee_id', ''),
            result.get('invoice_id') or 'N/A',
            result.get('year') or '',
            result.get('month') or '',
            result.get('approved_hours', ''),
            result.get('invoice_hours', ''),
            result.get('new_db_status', ''),
            result.get('row_count', ''),
        ]
        for c_idx, val in enumerate(row_data, 1):
            style_cell(ws3.cell(row=r_idx, column=c_idx, value=val), 'EAFAF1')

    for col, width in zip('ABCDEFGHIJK', [24, 16, 24, 16, 14, 8, 8, 15, 13, 16, 10]):
        ws3.column_dimensions[get_column_letter(ord(col) - 64)].width = width

    # ── Sheet 4: Unmatched / Ambiguous ───────────────────────────────────────
    # ── Sheet 4: Unmatched / Ambiguous ───────────────────────────────────────
    ws1 = wb.create_sheet('Unmatched')
    ws1.freeze_panes = 'A3'

    ws1.merge_cells('A1:L1')
    t = ws1['A1']
    t.value = f"Sync Comparison Report  |  Source: {source_filename}  |  {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    t.font, t.fill, t.alignment = Font(name='Arial', bold=True, size=12, color='FFFFFF'), fill('2C3E50'), center
    ws1.row_dimensions[1].height = 22

    cols1 = ['Status', 'Excel Name (From Timesheet)', 'Timesheet Employee ID', 'Year', 'Month', 'Row Count',
             'Hours (Approved)', 'Hours (Pending)', 'Hours (Other)', 'Total Hours', 'Possible DB Match', 'Possible DB Employee IDs']
    write_header(ws1, 2, cols1, 'C0392B')

    status_colors = {'UNMATCHED': 'FADBD8', 'AMBIGUOUS': 'FDEBD0', 'PERIOD_MISMATCH': 'FEF9E7'}

    for r_idx, result in enumerate(unmatched_results, start=3):
        status = result.get('status', '')
        fc     = status_colors.get(status, 'FFFFFF')

        # Get the group of rows for this person
        key = tuple(result.get('group_key') or (
            result.get('excel_first', ''), result.get('excel_last', ''),
            result.get('year', 0), result.get('month', 0)
        ))
        group = groups.get(key, [])
        
        # Calculate hours by approval status
        hours_by_status = {}
        for row in group:
            approval = (_get_col(row, COL_APPROVAL) or 'Other').strip().lower()
            if approval == 'approved':
                approval_key = 'Approved'
            elif approval == 'pending':
                approval_key = 'Pending'
            else:
                approval_key = 'Other'
            
            hours = _to_float((_get_col(row, COL_HOURS) or _get_col(row, 'hours') or '0').strip())
            hours_by_status[approval_key] = hours_by_status.get(approval_key, 0) + hours
        
        hours_approved = hours_by_status.get('Approved', 0)
        hours_pending = hours_by_status.get('Pending', 0)
        hours_other = hours_by_status.get('Other', 0)
        total_hours = hours_approved + hours_pending + hours_other

        possible = _possible_db_matches_for_report(result, group, db_invoices)

        row_data = [
            status,
            result.get('excel_name', ''),
            result.get('employee_id', ''),
            result.get('year') or '',
            result.get('month') or '',
            result.get('row_count', ''),
            f'{hours_approved:.2f}' if hours_approved > 0 else '',
            f'{hours_pending:.2f}' if hours_pending > 0 else '',
            f'{hours_other:.2f}' if hours_other > 0 else '',
            f'{total_hours:.2f}',
            ', '.join([p.get("resource_name") for p in possible if p.get("resource_name")]) if possible else '— no db match —',
            ', '.join([str(p.get("employee_id") or "") for p in possible if p.get("employee_id")]) if possible else '',
        ]
        for c_idx, val in enumerate(row_data, 1):
            style_cell(ws1.cell(row=r_idx, column=c_idx, value=val), fc)

    for col, width in zip('ABCDEFGHIJKL', [14, 26, 16, 8, 8, 10, 15, 15, 13, 12, 28, 22]):
        ws1.column_dimensions[get_column_letter(ord(col) - 64)].width = width
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _upload_comparison_report(unmatched_results: list, all_results: list, db_invoices: list, source_filename: str, groups: dict):
    try:
        report_bytes = _generate_comparison_report(unmatched_results, all_results, db_invoices, source_filename, groups)
        m = re.search(r"timesheet_(\d{4})_(\d{2})\.", str(source_filename or ""), re.I)
        if m:
            report_name = f"sync_report_{m.group(1)}_{m.group(2)}.xlsx"
        else:
            ts = datetime.now().strftime('%Y%m%d_%H%M')
            report_name = f"sync_report_{ts}.xlsx"

        url = _helpers().upload_sync_report_to_sharepoint(
            report_bytes, report_name
        )

        if url:
            logger.info("Comparison report uploaded: %s → %s", report_name, url)
        else:
            logger.warning("Comparison report upload returned no URL for: %s", report_name)
    except Exception as e:
        logger.error("Comparison report upload failed: %s", e)


def _run_standalone_employee_id_hours_auto_approve():
    """
    Existing data remediation (runs every sync-excel request, own DB connection):
    Auto-approve when employee_id is set, approved_hours matches invoice_hours,
    and status is Pending or Need Approval.
    """
    conn = _helpers().get_sql_connection()
    cur = conn.cursor()
    try:
        cur.execute("""
            UPDATE invoices
            SET approval_status = 'Approved',
                status = 'Approved',
                excel_updated_at = NOW(),
                last_updated_at = NOW()
            WHERE TRIM(COALESCE(employee_id, '')) <> ''
              AND approved_hours IS NOT NULL
              AND invoice_hours IS NOT NULL
              AND ABS(approved_hours::numeric - invoice_hours::numeric) <= 0.01
              AND LOWER(TRIM(COALESCE(approval_status, 'pending'))) IN ('pending', 'need approval')
        """)
        changed = cur.rowcount or 0
        conn.commit()
        if changed:
            logger.info("Auto-approved %s invoice rows (employee_id + exact hours match).", changed)
    finally:
        cur.close()
        conn.close()


def _possible_db_matches_for_report(result: dict, group: list, db_invoices: list) -> list:
    yr = int(result.get('year') or 0)
    mo = int(result.get('month') or 0)
    month_invoices = [inv for inv in db_invoices if _pay_period_matches(inv, yr, mo)] if yr and mo else list(db_invoices)
    scope_invoices = list(db_invoices) if result.get("status") == "PERIOD_MISMATCH" else month_invoices

    emp_id = (result.get("employee_id") or _employee_id_from_group(group) or "").strip().lower()
    if emp_id:
        by_emp = [
            {
                "resource_name": inv.get("resource_name"),
                "employee_id": inv.get("employee_id"),
            }
            for inv in scope_invoices
            if str(inv.get("employee_id") or "").strip().lower() == emp_id
        ]
        if by_emp:
            return by_emp

    # Fallback only when employee id is missing/unresolved
    excel_name = (result.get('excel_name') or "").strip()
    if not excel_name:
        excel_name = f"{result.get('excel_first', '')} {result.get('excel_last', '')}".strip()
    out = []
    for inv in scope_invoices:
        dbn = inv.get("resource_name") or ""
        if not dbn:
            continue
        if _human_name_gate_match(excel_name, dbn):
            out.append({
                "resource_name": dbn,
                "employee_id": inv.get("employee_id"),
            })
    return out

from decimal import Decimal

def _to_float(val) -> float:
    """Safely convert DB numeric/Decimal/string to float."""
    if val is None:
        return 0.0
    if isinstance(val, Decimal):
        return float(val)
    try:
        return float(str(val).strip())
    except (ValueError, TypeError):
        return 0.0